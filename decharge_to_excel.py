"""
decharge_to_excel.py
--------------------
1. Extrait les DUM N° et Qté BRUT depuis un PDF Décharge.
2. Retrouve chaque DUM N° dans le fichier Excel AT (colonne A).
3. Appende  -<qte_brut>  à la formule existante de la colonne QTE.REST (col I).
4. Sauvegarde le fichier modifié en .xlsx.

Usage:
    python decharge_to_excel.py <fichier_pdf> <fichier_excel>

Le fichier Excel de sortie sera <nom_original>_MAJ.xlsx
"""

import sys
import re
from pathlib import Path
import pdfplumber
import openpyxl
import win32com.client


# ─────────────────────────────────────────────
# 1.  Extraction PDF
# ─────────────────────────────────────────────

# Bornes horizontales de la colonne "Qté BRUT" (en points PDF)
# D'après inspection : en-tête à x0≈605, valeurs entre x0≈610 et x1≈690
QTE_BRUT_X_MIN = 610
QTE_BRUT_X_MAX = 690

# Patron d'un DUM N° : commence par des chiffres puis points/underscores
DUM_PATTERN = re.compile(r'^\d{3}\.\d{2}\.\d{2}\.\d{5}[_\w]+$')


def parse_french_number(text: str) -> float:
    """Convertit un nombre au format français (espace milliers, virgule décimale) en float.

    Gère aussi l'artefact d'encodage PDF où chaque caractère est doublé
    (ex : '1199,,1100' → '19,10').
    """
    # Artefact PDF : tous les caractères sont doublés (ex: '1199,,1100' → '19,10')
    stripped = text.replace(' ', '')
    if len(stripped) >= 2 and len(stripped) % 2 == 0:
        if all(stripped[i] == stripped[i + 1] for i in range(0, len(stripped), 2)):
            stripped = stripped[::2]

    cleaned = stripped.replace(',', '.')
    return float(cleaned)


def group_words_by_row(words: list, tolerance: float = 3.0) -> list[list[dict]]:
    """Regroupe les mots de pdfplumber par ligne (coordonnée 'top' similaire)."""
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: w['top'])
    rows = []
    current_row = [sorted_words[0]]
    for word in sorted_words[1:]:
        if abs(word['top'] - current_row[0]['top']) <= tolerance:
            current_row.append(word)
        else:
            rows.append(current_row)
            current_row = [word]
    rows.append(current_row)
    return rows


def extract_pdf_data(pdf_path: str) -> list[tuple[str, float]]:
    """
    Retourne une liste de (dum_numero, qte_brut) pour chaque ligne produit du PDF.
    """
    results = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            rows = group_words_by_row(words)
            for row_words in rows:
                # Chercher un DUM N° dans cette ligne
                dum = None
                for w in row_words:
                    if DUM_PATTERN.match(w['text']):
                        dum = w['text']
                        break
                if dum is None:
                    continue

                # Collecter les mots de la colonne Qté BRUT, triés par x0
                brut_words = sorted(
                    [w for w in row_words if QTE_BRUT_X_MIN <= w['x0'] <= QTE_BRUT_X_MAX],
                    key=lambda w: w['x0']
                )
                if not brut_words:
                    continue

                raw = ' '.join(w['text'] for w in brut_words)
                try:
                    qte_brut = parse_french_number(raw)
                except ValueError:
                    print(f"  ⚠ Impossible de parser Qté BRUT '{raw}' pour {dum}")
                    continue

                results.append((dum, qte_brut))
    return results


# ─────────────────────────────────────────────
# 2.  Normalisation DUM N° pour la comparaison
# ─────────────────────────────────────────────

def normalize_dum(dum: str) -> str:
    """
    Normalise un DUM N° pour la comparaison.
    PDF  : '309.22.24.03241_S5'  →  '309,22,24,03241 s5'
    Excel: '309,22,24,03241 S5'  →  '309,22,24,03241 s5'
    """
    return dum.replace('.', ',').replace('_', ' ').lower()


# ─────────────────────────────────────────────
# 3.  Mise à jour du fichier Excel
# ─────────────────────────────────────────────

QTE_REST_COL = 9   # colonne I (1-based)
DUM_COL      = 1   # colonne A (1-based)


def format_number(value: float) -> str:
    """
    Retourne une représentation compacte du nombre pour l'insérer dans une formule Excel.
    Ex: 280.0 → '280', 2.2 → '2.2', 5043.0 → '5043'
    """
    if value == int(value):
        return str(int(value))
    # Supprime les zéros inutiles
    return f'{value:g}'


def update_excel(excel_path: str, pdf_data: list[tuple[str, float]], output_path: str):
    """
    Charge le fichier Excel, met à jour les formules QTE.REST dans les 2 premiers
    onglets et sauvegarde.
    """
    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.worksheets

    updated = []
    not_found = []

    for dum_pdf, qte_brut in pdf_data:
        key = normalize_dum(dum_pdf)
        subtraction = format_number(qte_brut)
        found_anywhere = False

        for ws in sheets[:2]:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cell_val = row[DUM_COL - 1].value
                if not (cell_val and isinstance(cell_val, str)):
                    continue
                if normalize_dum(cell_val) != key:
                    continue

                # Correspondance trouvée dans cet onglet
                found_anywhere = True
                cell = ws.cell(row=row[0].row, column=QTE_REST_COL)
                current = cell.value

                if isinstance(current, str) and current.startswith('='):
                    new_formula = f'{current}-{subtraction}'
                elif isinstance(current, (int, float)):
                    new_formula = f'={current}-{subtraction}'
                else:
                    new_formula = f'=-{subtraction}'

                cell.value = new_formula
                updated.append((dum_pdf, ws.title, row[0].row, current, new_formula))
                break  # Stop after first match in this sheet

        if not found_anywhere:
            not_found.append(dum_pdf)

    # ── Rapport ──────────────────────────────
    print(f"\n{'='*60}")
    print(f"  RAPPORT DE MISE À JOUR")
    print(f"{'='*60}")
    print(f"\n  ✅ {len(updated)} ligne(s) mise(s) à jour :\n")
    for dum, sheet, row_num, old, new in updated:
        print(f"    • {dum}  (onglet '{sheet}', ligne {row_num})")
        print(f"      Avant : {old}")
        print(f"      Après : {new}\n")

    if not_found:
        print(f"  ⚠  {len(not_found)} DUM N° non trouvé(s) dans les 2 onglets :")
        for d in not_found:
            print(f"     - {d}")

    wb.save(output_path)
    print(f"\n  💾 Fichier sauvegardé : {output_path}")
    print(f"{'='*60}\n")


# ─────────────────────────────────────────────
# 4.  Point d'entrée
# ─────────────────────────────────────────────

def convert_xls_to_xlsx(xls_path: str) -> str:
    """Convertit un .xls en .xlsx via Excel COM (préserve la mise en forme)."""
    p = Path(xls_path)
    if p.suffix.lower() == '.xlsx':
        return xls_path
    xlsx_path = p.with_suffix('.xlsx')
    print(f"  🔄 Conversion {p.name} → {xlsx_path.name} via Excel...")
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(p.resolve()))
        wb.SaveAs(str(xlsx_path.resolve()), FileFormat=51)  # 51 = xlOpenXMLWorkbook
        wb.Close(False)
    finally:
        excel.Quit()
    return str(xlsx_path)


def main():
    base_dir    = Path(__file__).parent
    pdf_dir     = base_dir / 'input_pdf'
    excel_dir   = base_dir / 'excel'
    output_dir  = base_dir / 'output'
    output_dir.mkdir(exist_ok=True)

    # Résolution des chemins : arguments CLI prioritaires, sinon auto-détection
    if len(sys.argv) >= 3:
        pdf_path   = sys.argv[1]
        excel_path = sys.argv[2]
    else:
        pdfs = list(pdf_dir.glob('*.pdf'))
        excels = list(excel_dir.glob('*.xls*'))

        if not pdfs:
            print(f"  ❌ Aucun PDF trouvé dans {pdf_dir}")
            sys.exit(1)
        if not excels:
            print(f"  ❌ Aucun fichier Excel trouvé dans {excel_dir}")
            sys.exit(1)

        if len(pdfs) > 1:
            print("  Plusieurs PDF trouvés, sélectionnez :")
            for i, f in enumerate(pdfs):
                print(f"    [{i}] {f.name}")
            idx = int(input("  Numéro : "))
            pdf_path = str(pdfs[idx])
        else:
            pdf_path = str(pdfs[0])

        if len(excels) > 1:
            print("  Plusieurs fichiers Excel trouvés, sélectionnez :")
            for i, f in enumerate(excels):
                print(f"    [{i}] {f.name}")
            idx = int(input("  Numéro : "))
            excel_path = str(excels[idx])
        else:
            excel_path = str(excels[0])

    # Sortie : dossier output/, suffixe _MAJ.xlsx
    p = Path(excel_path)
    output_path = str(output_dir / (p.stem + '_MAJ.xlsx'))

    print(f"\n📄 PDF    : {pdf_path}")
    print(f"📊 Excel  : {excel_path}")
    print(f"💾 Sortie : {output_path}\n")

    # Conversion XLS → XLSX si nécessaire
    xlsx_input = convert_xls_to_xlsx(excel_path)

    # Extraction PDF
    print("  🔍 Extraction des données PDF...")
    pdf_data = extract_pdf_data(pdf_path)
    if not pdf_data:
        print("  ❌ Aucune donnée extraite du PDF. Vérifiez le fichier.")
        sys.exit(1)

    print(f"\n  Données extraites du PDF ({len(pdf_data)} produit(s)) :")
    print(f"  {'DUM N°':<30} {'Qté BRUT':>12}")
    print(f"  {'-'*43}")
    for dum, qte in pdf_data:
        print(f"  {dum:<30} {qte:>12g}")

    # Mise à jour Excel
    update_excel(xlsx_input, pdf_data, output_path)


if __name__ == '__main__':
    main()
