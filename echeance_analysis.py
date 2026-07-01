"""
echeance_analysis.py — Génère le tableau de suivi des échéances AT (in-memory).
"""

import ast
import io
import operator
import re
import datetime
from collections import OrderedDict

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ── Colour palette keyed on "vrais famille" ──────────────────────────────────
FAMILLE_COLORS = {
    "MDF":       "C4955A",
    "MDF MINCE": "DBBE94",
    "VERRE":     "2E8B57",
    "ACCESSOIR": "4A7FA5",
    "EMBALAGE":  "C4922A",
    "DIVER":     "7F8C8D",
    "BOIS":      "7B3F00",
    "PROBLEME":  "C0392B",
}
DEFAULT_COLOR = "95A5A6"

# ── Remorque percentages ──────────────────────────────────────────────────────
REMORQUE_CAPACITY = 33_000
FAMILLE_REMORQUE_PCT = {
    "ACCESSOIR": 2,
    "MDF MINCE": 25,
    "MDF":       45,
    "VERRE":     20,
    "EMBALAGE":   3,
    "DIVER":      5,
}

# ── Style constants ───────────────────────────────────────────────────────────
DEADLINE_COLOR = "000000"
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL   = PatternFill("solid", fgColor="2F5496")
ALT_ROW_FILL   = PatternFill("solid", fgColor="F2F2F2")
SUBHEAD_FONT   = Font(bold=True, color="FFFFFF", size=9)
DATA_FONT      = Font(size=9)
QTE_FONT       = Font(bold=True, color="FFFFFF", size=8)
THIN           = Side(style="thin", color="D9D9D9")
THIN_BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONTH_FR       = ["Jan","Fev","Mar","Avr","Mai","Juin",
                  "Juil","Aout","Sep","Oct","Nov","Dec"]


def _make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _month_iter(sy, sm, ey, em):
    y, m = sy, sm
    while (y, m) <= (ey, em):
        yield y, m
        m += 1
        if m > 12:
            m, y = 1, y + 1


def _col_for_month(start_col, sy, sm, y, m):
    return start_col + (y - sy) * 12 + (m - sm)


def _add_deadline_months(d, n):
    """Return date that is n months after d."""
    m = d.month + n
    y = d.year + (m - 1) // 12
    m = (m - 1) % 12 + 1
    return datetime.date(y, m, 1)


_ARITH_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def _safe_eval(expr: str):
    """Evaluate a pure-numeric arithmetic expression with no cell references."""
    def _node(n):
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return n.value
        if isinstance(n, ast.BinOp):
            op = _ARITH_OPS.get(type(n.op))
            if op:
                return op(_node(n.left), _node(n.right))
        if isinstance(n, ast.UnaryOp):
            op = _ARITH_OPS.get(type(n.op))
            if op:
                return op(_node(n.operand))
        raise ValueError(f"unsupported: {n}")
    try:
        return _node(ast.parse(expr, mode="eval").body)
    except Exception:
        return None


def _resolve_cell(ws, row: int, col: int, _seen: set | None = None):
    """Return the numeric value of a cell, evaluating simple same-sheet formulas."""
    if _seen is None:
        _seen = set()
    key = (row, col)
    if key in _seen:
        return None
    _seen.add(key)

    val = ws.cell(row, col).value
    if val is None:
        return None
    if not isinstance(val, str) or not val.startswith("="):
        return val if isinstance(val, (int, float, datetime.datetime, datetime.date)) else None

    # Strip leading = and optional leading +
    expr = val[1:].lstrip("+")

    # Replace all A1-style same-sheet references (no sheet prefix) with their numeric values
    # Pass a copy of _seen so sibling resolutions don't pollute each other's cycle check
    def _sub_ref(m):
        ref_col = column_index_from_string(m.group(1).upper())
        ref_row = int(m.group(2))
        v = _resolve_cell(ws, ref_row, ref_col, set(_seen))
        if isinstance(v, (int, float)):
            return str(v)
        return "0"

    expr = re.sub(r"(?<!['\w])([A-Za-z]+)(\d+)", _sub_ref, expr)
    return _safe_eval(expr)


def generate_echeance_excel(excel_bytes: bytes, analysis_date: datetime.date) -> tuple[bytes, dict]:
    """
    Read sheet 2 of the uploaded AT Excel file, build the écheances tableau,
    and return (xlsx_bytes, summary_dict).
    """
    # ── Read source data (sheet index 1) — formulas kept so we can evaluate them
    wb_src = load_workbook(io.BytesIO(excel_bytes), data_only=False)
    ws_src = wb_src.worksheets[1]

    # Column positions (1-based): A=DUM, E=vrais famille, G=échéance, U=P.NE KG restant
    COL_DUM_SRC, COL_FAM_SRC, COL_ECH_SRC, COL_PNE_SRC = 1, 5, 7, 21

    rows = []
    for xl_row in ws_src.iter_rows(min_row=3):
        dum_cell = ws_src.cell(xl_row[0].row, COL_DUM_SRC).value
        if dum_cell is None:
            continue
        dum = str(dum_cell).strip()

        fam_val = ws_src.cell(xl_row[0].row, COL_FAM_SRC).value
        vrais_fam = str(fam_val).strip() if fam_val else ""

        echeance = ws_src.cell(xl_row[0].row, COL_ECH_SRC).value
        if isinstance(echeance, str) and echeance.startswith("="):
            echeance = _resolve_cell(ws_src, xl_row[0].row, COL_ECH_SRC)

        pne_rest = _resolve_cell(ws_src, xl_row[0].row, COL_PNE_SRC)

        if isinstance(echeance, datetime.datetime) and pne_rest is not None:
            rows.append({
                "dum":      dum,
                "famille":  vrais_fam,
                "echeance": echeance,
                "qte_rest": pne_rest,
            })

    rows.sort(key=lambda r: r["echeance"])

    # ── Month range: analysis_date → analysis_date + 24 months ───────────────
    START_YEAR, START_MONTH = analysis_date.year, analysis_date.month
    deadline = _add_deadline_months(analysis_date, 24)
    END_YEAR, END_MONTH     = deadline.year, deadline.month

    months       = list(_month_iter(START_YEAR, START_MONTH, END_YEAR, END_MONTH))
    DEADLINE_IDX = len(months) - 1

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TABLEAU ECHEANCES"

    FIXED_COLS  = 5
    MONTH_START = FIXED_COLS + 1
    COL_DUM, COL_FAM, COL_ECH, COL_QTE, COL_REM = 1, 2, 3, 4, 5
    total_cols = FIXED_COLS + len(months)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(1, 1,
        f"TABLEAU DE SUIVI DES ECHEANCES AT  MEDIDIS  "
        f"(analyse: {analysis_date.strftime('%d/%m/%Y')})")
    tc.fill      = HEADER_FILL
    tc.font      = Font(bold=True, color="FFFFFF", size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header row
    for c, h in enumerate(["N DOSSIER","FAMILLE","ECHEANCE","P. NE KG REST.","REMORQUES"], 1):
        cell = ws.cell(2, c, h)
        cell.fill      = SUBHEAD_FILL
        cell.font      = SUBHEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    for idx, (y, m) in enumerate(months):
        col  = MONTH_START + idx
        cell = ws.cell(2, col)
        cell.border = THIN_BORDER
        if idx == DEADLINE_IDX:
            cell.value     = "24M\nLIMITE"
            cell.fill      = _make_fill(DEADLINE_COLOR)
            cell.font      = Font(bold=True, color="FFFFFF", size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.value     = f"{MONTH_FR[m-1]}\n{str(y)[2:]}"
            cell.fill      = SUBHEAD_FILL
            cell.font      = SUBHEAD_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Data rows
    for r_idx, r in enumerate(rows):
        row_num  = r_idx + 3
        alt      = (r_idx % 2 == 1)
        bg       = ALT_ROW_FILL if alt else None
        bar_fill = _make_fill(FAMILLE_COLORS.get(r["famille"], DEFAULT_COLOR))

        def write_fixed(col, value=None, number_format=None, align="left"):
            cell = ws.cell(row_num, col, value)
            cell.font      = DATA_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if bg:
                cell.fill = bg
            if number_format:
                cell.number_format = number_format
            return cell

        write_fixed(COL_DUM, r["dum"])
        write_fixed(COL_FAM, r["famille"], align="center")
        write_fixed(COL_ECH, r["echeance"], number_format="DD/MM/YYYY", align="center")
        qc = write_fixed(COL_QTE, r["qte_rest"], align="right")
        qc.number_format = '#,##0.00'

        pct = FAMILLE_REMORQUE_PCT.get(r["famille"])
        rc  = write_fixed(COL_REM, align="center")
        if pct is not None:
            rc.value         = f"={get_column_letter(COL_QTE)}{row_num}/({pct/100}*33000)"
            rc.number_format = '#,##0.00'
        else:
            rc.value = "—"

        ech_col = _col_for_month(MONTH_START, START_YEAR, START_MONTH,
                                 r["echeance"].year, r["echeance"].month)

        for idx, (y, m) in enumerate(months):
            col  = MONTH_START + idx
            cell = ws.cell(row_num, col)
            cell.border = THIN_BORDER
            if idx == DEADLINE_IDX:
                cell.fill      = _make_fill(DEADLINE_COLOR)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col <= ech_col:
                cell.fill = bar_fill
                if col == ech_col:
                    cell.value         = r["qte_rest"]
                    cell.number_format = '#,##0.##'
                    cell.font          = QTE_FONT
                    cell.alignment     = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if bg:
                    cell.fill = bg

        ws.row_dimensions[row_num].height = 18

    # Column widths
    ws.column_dimensions[get_column_letter(COL_DUM)].width = 24
    ws.column_dimensions[get_column_letter(COL_FAM)].width = 16
    ws.column_dimensions[get_column_letter(COL_ECH)].width = 13
    ws.column_dimensions[get_column_letter(COL_QTE)].width = 14
    ws.column_dimensions[get_column_letter(COL_REM)].width = 12
    for idx in range(len(months)):
        ltr = get_column_letter(MONTH_START + idx)
        ws.column_dimensions[ltr].width = 7 if idx == DEADLINE_IDX else 10

    ws.freeze_panes = ws.cell(3, MONTH_START)
    last_data_row = 2 + len(rows)
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{last_data_row}"

    # ── Sheet 2: Totals per famille ───────────────────────────────────────────
    ws2 = wb.create_sheet("TOTAUX PAR FAMILLE")

    seen = {r["famille"] for r in rows}
    familles = [f for f in FAMILLE_COLORS if f in seen]
    for f in seen:
        if f not in familles:
            familles.append(f)

    MAIN         = "TABLEAU ECHEANCES"
    fam_col_main = get_column_letter(COL_FAM)
    qte_col_main = get_column_letter(COL_QTE)

    ws2.merge_cells("A1:D1")
    t2 = ws2.cell(1, 1, "TOTAUX PAR FAMILLE")
    t2.fill      = HEADER_FILL
    t2.font      = Font(bold=True, color="FFFFFF", size=12)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    for col, h in enumerate(["FAMILLE","P. NE KG REST.","% REMORQUE","NB REMORQUES"], 1):
        cell = ws2.cell(2, col, h)
        cell.fill      = SUBHEAD_FILL
        cell.font      = SUBHEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws2.row_dimensions[2].height = 28

    first_data = 3
    last_data  = first_data + len(familles) - 1

    for r_idx, fam in enumerate(familles):
        row_num  = first_data + r_idx
        pct      = FAMILLE_REMORQUE_PCT.get(fam)
        bar_fill = _make_fill(FAMILLE_COLORS.get(fam, DEFAULT_COLOR))
        wfont    = Font(color="FFFFFF", size=9)
        bfont    = Font(bold=True, color="FFFFFF", size=9)

        def w2(col, value, fmt=None, align="center"):
            cell = ws2.cell(row_num, col, value)
            cell.fill      = bar_fill
            cell.font      = wfont
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if fmt:
                cell.number_format = fmt
            return cell

        fc = w2(1, fam)
        fc.font = bfont

        kc = w2(2,
            f"=SUMIF('{MAIN}'!${fam_col_main}:${fam_col_main},"
            f"A{row_num},'{MAIN}'!${qte_col_main}:${qte_col_main})",
            fmt='#,##0.00', align="right")

        if pct is not None:
            pc = w2(3, pct / 100, fmt='0%')
        else:
            pc = w2(3, "—")

        if pct is not None:
            rc = w2(4, f"=B{row_num}/(C{row_num}*33000)", fmt='#,##0.00')
            rc.font = bfont
        else:
            rc = w2(4, "—")
            rc.font = bfont

        ws2.row_dimensions[row_num].height = 20

    # Total row
    total_row = last_data + 1
    for col in range(1, 5):
        cell = ws2.cell(total_row, col)
        cell.fill      = HEADER_FILL
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.cell(total_row, 1, "TOTAL")
    gtc = ws2.cell(total_row, 2, f"=SUM(B{first_data}:B{last_data})")
    gtc.number_format = '#,##0.00'
    gtc.alignment     = Alignment(horizontal="right", vertical="center")
    ws2.cell(total_row, 3, "")
    trc = ws2.cell(total_row, 4, f"=B{total_row}/33000")
    trc.number_format = '#,##0.00'
    trc.alignment     = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[total_row].height = 20

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)

    dl_y, dl_m = months[DEADLINE_IDX]
    summary = {
        "nb_dossiers":   len(rows),
        "start_label":   f"{MONTH_FR[START_MONTH-1]}-{str(START_YEAR)[2:]}",
        "deadline_label": f"{MONTH_FR[dl_m-1]}-{str(dl_y)[2:]}",
        "familles": familles,
    }
    return buf.getvalue(), summary
