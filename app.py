"""
app.py — Interface Streamlit pour décharge + tableau écheances AT
"""

import io
import re
import datetime
import tempfile
from pathlib import Path

import pdfplumber
import openpyxl
import pandas as pd
import streamlit as st

from echeance_analysis import generate_echeance_excel, FAMILLE_COLORS, FAMILLE_REMORQUE_PCT

# ─────────────────────────────────────────────
# Logique métier — Décharge (Tab 1)
# ─────────────────────────────────────────────

QTE_BRUT_X_MIN = 610
QTE_BRUT_X_MAX = 690
DUM_PATTERN = re.compile(r'^\d{3}\.\d{2}\.\d{2}\.\d{5}[_\w]+$')
QTE_REST_COL = 9
DUM_COL = 1


def parse_french_number(text: str) -> float:
    stripped = text.replace(' ', '')
    if len(stripped) >= 2 and len(stripped) % 2 == 0:
        if all(stripped[i] == stripped[i + 1] for i in range(0, len(stripped), 2)):
            stripped = stripped[::2]
    cleaned = stripped.replace(',', '.')
    return float(cleaned)


def group_words_by_row(words: list, tolerance: float = 3.0) -> list:
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: w['top'])
    rows = []
    current_row = [sorted_words[0]]
    for word in sorted_words[1:]:
        if abs(word['top'] - current_row[0]['top']) <= tolerance:
            current_row.append(word)
        else:
            rows.append(current_row)
            current_row = [word]
    rows.append(current_row)
    return rows


def extract_pdf_data(pdf_bytes: bytes) -> tuple[list, list]:
    results = []
    warnings = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            rows = group_words_by_row(words)
            for row_words in rows:
                dum = None
                for w in row_words:
                    if DUM_PATTERN.match(w['text']):
                        dum = w['text']
                        break
                if dum is None:
                    continue
                brut_words = sorted(
                    [w for w in row_words if QTE_BRUT_X_MIN <= w['x0'] <= QTE_BRUT_X_MAX],
                    key=lambda w: w['x0']
                )
                if not brut_words:
                    continue
                raw = ' '.join(w['text'] for w in brut_words)
                try:
                    qte_brut = parse_french_number(raw)
                except ValueError:
                    warnings.append(f"Impossible de parser Qté BRUT '{raw}' pour {dum}")
                    continue
                results.append((dum, qte_brut))
    return results, warnings


def normalize_dum(dum: str) -> str:
    return dum.replace('.', ',').replace('_', ' ').lower()


def format_number(value: float) -> str:
    if value == int(value):
        return str(int(value))
    return f'{value:g}'


def find_qte_rest_col(ws, fallback=QTE_REST_COL):
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'QTE.REST' in cell.value.upper():
                return cell.column
    return fallback


def update_excel(excel_bytes: bytes, pdf_data: list) -> tuple[bytes, list, list]:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheets = wb.worksheets
    updated = []
    not_found = []

    for dum_pdf, qte_brut in pdf_data:
        key = normalize_dum(dum_pdf)
        subtraction = format_number(qte_brut)
        found_anywhere = False

        for ws in sheets[:2]:
            qte_rest_col = find_qte_rest_col(ws)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cell_val = row[DUM_COL - 1].value
                if not (cell_val and isinstance(cell_val, str)):
                    continue
                if normalize_dum(cell_val) != key:
                    continue
                found_anywhere = True
                cell = ws.cell(row=row[0].row, column=qte_rest_col)
                current = cell.value
                if isinstance(current, str) and current.startswith('='):
                    new_formula = f'{current}-{subtraction}'
                elif isinstance(current, (int, float)):
                    new_formula = f'={current}-{subtraction}'
                else:
                    new_formula = f'=-{subtraction}'
                cell.value = new_formula
                updated.append((dum_pdf, ws.title, row[0].row, current, new_formula))
                break

        if not found_anywhere:
            not_found.append(dum_pdf)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), updated, not_found


def convert_xls_to_xlsx_bytes(xls_bytes: bytes, original_name: str) -> bytes:
    try:
        import win32com.client
    except ImportError:
        raise RuntimeError(
            "La conversion .xls → .xlsx via Excel COM n'est disponible que sous Windows. "
            "Veuillez convertir votre fichier en .xlsx avant de l'importer."
        )
    with tempfile.TemporaryDirectory() as tmp:
        xls_path = Path(tmp) / original_name
        xlsx_path = xls_path.with_suffix('.xlsx')
        xls_path.write_bytes(xls_bytes)
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(xls_path.resolve()))
            wb.SaveAs(str(xlsx_path.resolve()), FileFormat=51)
            wb.Close(False)
        finally:
            excel.Quit()
        return xlsx_path.read_bytes()


# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────

st.set_page_config(page_title="AT Medidis", page_icon="📊", layout="wide")
st.title("📊 AT Medidis — Outils d'analyse")

tab1, tab2 = st.tabs(["🔄 Mise à jour décharge", "📅 Tableau suivi échéances"])


# ─────────────────────────────────────────────
# Tab 1 — Mise à jour décharge
# ─────────────────────────────────────────────

with tab1:
    st.subheader("Mise à jour tableau de décharge")
    st.markdown("Importez le fichier Excel AT et le PDF Décharge pour mettre à jour les formules QTE.REST.")

    col1, col2 = st.columns(2)
    with col1:
        excel_file = st.file_uploader("Fichier Excel (.xls / .xlsx)", type=["xls", "xlsx"],
                                      key="tab1_excel")
    with col2:
        pdf_file = st.file_uploader("PDF Décharge", type=["pdf"], key="tab1_pdf")

    if excel_file and pdf_file:
        if st.button("Lancer la mise à jour", type="primary", key="tab1_run"):
            with st.spinner("Traitement en cours…"):
                pdf_bytes   = pdf_file.read()
                excel_bytes = excel_file.read()
                excel_name  = excel_file.name

                if excel_name.lower().endswith('.xls'):
                    try:
                        with st.spinner("Conversion .xls → .xlsx via Excel…"):
                            excel_bytes = convert_xls_to_xlsx_bytes(excel_bytes, excel_name)
                        excel_name = Path(excel_name).stem + '.xlsx'
                    except Exception as e:
                        st.error(f"Échec de la conversion XLS : {e}")
                        st.stop()

                pdf_data, pdf_warnings = extract_pdf_data(pdf_bytes)

                if pdf_warnings:
                    for w in pdf_warnings:
                        st.warning(w)

                if not pdf_data:
                    st.error("Aucune donnée extraite du PDF. Vérifiez le fichier.")
                    st.stop()

                st.subheader(f"Données extraites du PDF — {len(pdf_data)} produit(s)")
                df = pd.DataFrame(pdf_data, columns=["DUM N°", "Qté BRUT"])
                st.dataframe(df, use_container_width=True, hide_index=True)

                result_bytes, updated, not_found = update_excel(excel_bytes, pdf_data)

                st.subheader("Rapport")
                if updated:
                    st.success(f"{len(updated)} ligne(s) mise(s) à jour")
                    rows_report = []
                    for dum, sheet, row_num, old, new in updated:
                        rows_report.append({"DUM N°": dum, "Onglet": sheet,
                                            "Ligne": row_num, "Avant": old, "Après": new})
                    st.dataframe(pd.DataFrame(rows_report), use_container_width=True, hide_index=True)
                else:
                    st.warning("Aucune ligne mise à jour.")

                if not_found:
                    st.error(f"{len(not_found)} DUM N° non trouvé(s) dans les 2 premiers onglets :")
                    for d in not_found:
                        st.markdown(f"- `{d}`")

                output_name = Path(excel_name).stem + '_MAJ.xlsx'
                st.download_button(
                    label="⬇️ Télécharger le fichier mis à jour",
                    data=result_bytes,
                    file_name=output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    else:
        st.info("Veuillez importer les deux fichiers pour continuer.")


# ─────────────────────────────────────────────
# Tab 2 — Tableau suivi échéances AT
# ─────────────────────────────────────────────

with tab2:
    st.subheader("Tableau de suivi des échéances AT")
    st.markdown(
        "Importez le fichier AT Excel (onglet 2 : *MAJ-ECH compte*). "
        "L'analyse génère un tableau Gantt par dossier + le récapitulatif des remorques par famille."
    )

    col_a, col_b = st.columns([2, 1])
    with col_a:
        at_file = st.file_uploader("Fichier AT Excel (.xlsx)", type=["xlsx"], key="tab2_excel")
    with col_b:
        analysis_date = st.date_input(
            "Date d'analyse",
            value=datetime.date.today(),
            help="La colonne noire (limite 24 mois) sera placée 24 mois après cette date.",
            key="tab2_date",
        )

    # Legend
    with st.expander("Légende couleurs & paramètres remorques"):
        legend_rows = []
        for fam, color in FAMILLE_COLORS.items():
            pct = FAMILLE_REMORQUE_PCT.get(fam)
            legend_rows.append({
                "Famille":        fam,
                "% remorque":     f"{pct}%" if pct else "—",
                "KG / remorque":  f"{int(pct/100*33000):,}" if pct else "—",
            })
        st.dataframe(pd.DataFrame(legend_rows), use_container_width=True, hide_index=True)
        st.caption("Capacité totale d'une remorque : 33 000 kg")

    if at_file:
        if st.button("Générer le tableau", type="primary", key="tab2_run"):
            with st.spinner("Génération en cours…"):
                try:
                    result_bytes, summary = generate_echeance_excel(
                        at_file.read(), analysis_date
                    )
                except Exception as e:
                    st.error(f"Erreur lors de la génération : {e}")
                    st.stop()

            # Summary metrics
            st.success("Tableau généré avec succès !")
            m1, m2, m3 = st.columns(3)
            m1.metric("Dossiers", summary["nb_dossiers"])
            m2.metric("Période", f"{summary['start_label']} → {summary['deadline_label']}")
            m3.metric("Limite 24 mois", summary["deadline_label"])

            output_name = Path(at_file.name).stem + '_ECHEANCES.xlsx'
            st.download_button(
                label="⬇️ Télécharger le tableau des échéances",
                data=result_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Importez le fichier AT Excel pour générer le tableau.")
